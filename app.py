"""
Choice Restaurant Parser — Flask Backend
Uses: Places API (New) — places.googleapis.com/v1
Run: python app.py  →  http://localhost:5000
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
import requests as req
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re, time, os, io
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__, static_folder=".")

API_KEY = os.environ.get("API_KEY", "")

BASE_URL = "https://places.googleapis.com/v1"

FIELD_MASK = (
    "places.id,"
    "places.displayName,"
    "places.formattedAddress,"
    "places.internationalPhoneNumber,"
    "places.websiteUri,"
    "places.rating,"
    "places.userRatingCount,"
    "places.types,"
    "places.googleMapsUri"
)

def api_headers():
    return {
        "X-Goog-Api-Key": API_KEY,
        "X-Goog-FieldMask": FIELD_MASK,
        "Content-Type": "application/json",
    }

# ─── CITY GRID ────────────────────────────────────────────────────────────────

CITY_GRIDS = {
    "warsaw":    {"lat": 52.237, "lng": 21.017, "r": 0.12, "step": 0.05},
    "kyiv":      {"lat": 50.450, "lng": 30.523, "r": 0.12, "step": 0.05},
    "prague":    {"lat": 50.075, "lng": 14.437, "r": 0.10, "step": 0.05},
    "krakow":    {"lat": 50.061, "lng": 19.937, "r": 0.08, "step": 0.04},
    "kraków":    {"lat": 50.061, "lng": 19.937, "r": 0.08, "step": 0.04},
    "riga":      {"lat": 56.946, "lng": 24.106, "r": 0.08, "step": 0.04},
    "tallinn":   {"lat": 59.437, "lng": 24.745, "r": 0.07, "step": 0.04},
    "vilnius":   {"lat": 54.687, "lng": 25.279, "r": 0.08, "step": 0.04},
    "bucharest": {"lat": 44.426, "lng": 26.103, "r": 0.10, "step": 0.05},
    "budapest":  {"lat": 47.497, "lng": 19.040, "r": 0.10, "step": 0.05},
}

def get_grid_cells(city):
    g = CITY_GRIDS.get(city.lower().replace("ó", "o").replace("ő", "o"))
    if not g:
        return None, None
    cells = []
    lat = g["lat"] - g["r"]
    while lat <= g["lat"] + g["r"] + 0.001:
        lng = g["lng"] - g["r"]
        while lng <= g["lng"] + g["r"] + 0.001:
            cells.append((round(lat, 4), round(lng, 4)))
            lng += g["step"]
        lat += g["step"]
    radius_m = int(g["step"] * 111000 * 0.8)
    return cells, radius_m


# ─── PLACES API (NEW) ─────────────────────────────────────────────────────────

def text_search(query, city, page_token=None):
    url = f"{BASE_URL}/places:searchText"
    body = {"textQuery": f"{query} in {city}", "maxResultCount": 20}
    if page_token:
        body["pageToken"] = page_token
    r = req.post(url, headers=api_headers(), json=body, timeout=12)
    return r.json()


def nearby_search(lat, lng, radius_m, keyword):
    url = f"{BASE_URL}/places:searchNearby"
    body = {
        "textQuery": keyword,
        "maxResultCount": 20,
        "locationRestriction": {
            "circle": {
                "center": {"latitude": lat, "longitude": lng},
                "radius": float(radius_m),
            }
        },
    }
    r = req.post(url, headers=api_headers(), json=body, timeout=12)
    return r.json()


def parse_place(p):
    return {
        "_id":     p.get("id", ""),
        "name":    p.get("displayName", {}).get("text", ""),
        "address": p.get("formattedAddress", ""),
        "phone":   p.get("internationalPhoneNumber", ""),
        "website": p.get("websiteUri", ""),
        "rating":  p.get("rating", ""),
        "reviews": p.get("userRatingCount", ""),
        "types":   p.get("types", []),
        "maps_url": p.get("googleMapsUri", ""),
    }


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def find_instagram(website_url):
    if not website_url:
        return ""
    try:
        r = req.get(website_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=5)
        matches = re.findall(
            r'https?://(?:www\.)?instagram\.com/([A-Za-z0-9._]+)(?:[/?"\']|$)', r.text
        )
        clean = [m for m in matches if not m.startswith(("p/", "reel", "explore", "sharer"))]
        if clean:
            return f"https://instagram.com/{clean[0]}"
    except Exception:
        pass
    return ""


IGNORE_TYPES = {"point_of_interest", "establishment", "food", "premise",
                "store", "locality", "political", "geocode"}
CUISINE_MAP = {
    "restaurant": "Restaurant", "cafe": "Cafe", "bar": "Bar",
    "bakery": "Bakery", "meal_takeaway": "Takeaway", "meal_delivery": "Delivery",
    "pizza_restaurant": "Pizza", "night_club": "Night Club", "brewery": "Brewery",
    "fast_food_restaurant": "Fast Food", "fine_dining_restaurant": "Fine Dining",
}

def get_cuisine(types_list):
    result = [CUISINE_MAP.get(t, t.replace("_", " ").title())
              for t in (types_list or []) if t not in IGNORE_TYPES]
    return ", ".join(result[:3])

def detect_platform(website):
    for p in ["choiceqr", "glovo", "bolt", "uber", "wolt", "pyszne", "lieferando"]:
        if p in (website or "").lower():
            return p.capitalize()
    return ""

def enrich(raw):
    website = raw["website"]
    return {
        "name":      raw["name"],
        "address":   raw["address"],
        "phone":     raw["phone"],
        "website":   website,
        "instagram": find_instagram(website),
        "rating":    raw["rating"],
        "reviews":   raw["reviews"],
        "cuisine":   get_cuisine(raw["types"]),
        "platform":  detect_platform(website),
        "maps_url":  raw["maps_url"],
    }


# ─── COLLECT ──────────────────────────────────────────────────────────────────

def collect_text_search(category, city):
    places = {}
    page_token = None
    for _ in range(3):
        data = text_search(category, city, page_token)
        if "error" in data:
            raise ValueError(f"Google API: {data['error'].get('message', 'error')}")
        for p in data.get("places", []):
            parsed = parse_place(p)
            if parsed["_id"]:
                places[parsed["_id"]] = parsed
        page_token = data.get("nextPageToken")
        if not page_token:
            break
        time.sleep(2)
    return places


def collect_grid(category, city):
    cells, radius_m = get_grid_cells(city)
    if not cells:
        return collect_text_search(category, city)
    places = {}
    for lat, lng in cells:
        data = nearby_search(lat, lng, radius_m, category)
        for p in data.get("places", []):
            parsed = parse_place(p)
            if parsed["_id"]:
                places[parsed["_id"]] = parsed
        time.sleep(0.4)
    return places


# ─── EXCEL ────────────────────────────────────────────────────────────────────

HEADERS = ["Name", "Address", "Phone", "Website", "Instagram",
           "Rating", "Reviews", "Cuisine", "Platform", "Google Maps"]

def build_excel(rows, city):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{city} Restaurants"
    ws.freeze_panes = "A2"
    col_widths = [28, 38, 16, 32, 32, 10, 10, 22, 20, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ri, row in enumerate(rows, 2):
        fill = "FFF3EF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if ci in (4, 5, 10) and val and str(val).startswith("http"):
                c.hyperlink = val
                c.font = Font(color="FF7855", underline="single")
        ws.row_dimensions[ri].height = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/search", methods=["POST"])
def search():
    data       = request.json
    city       = data.get("city", "").strip()
    categories = data.get("categories", ["restaurants"])
    grid_mode  = data.get("grid", False)

    if not city:
        return jsonify({"error": "Please enter a city"}), 400
    if not API_KEY:
        return jsonify({"error": "API_KEY not set in environment variables"}), 400
    if isinstance(categories, str):
        categories = [categories]

    try:
        all_places = {}
        for cat in categories:
            places = collect_grid(cat, city) if grid_mode else collect_text_search(cat, city)
            all_places.update(places)
    except ValueError as e:
        return jsonify({"error": str(e)}), 500

    results = []
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(enrich, p): p for p in all_places.values()}
        for future in as_completed(futures):
            r = future.result()
            if r and r["name"]:
                results.append(r)

    results.sort(key=lambda x: float(x["rating"] or 0), reverse=True)
    return jsonify({"count": len(results), "results": results})


@app.route("/export", methods=["POST"])
def export():
    data = request.json
    rows = data.get("rows", [])
    city = data.get("city", "export")
    excel_rows = [
        [r["name"], r["address"], r["phone"], r["website"], r["instagram"],
         r["rating"], r["reviews"], r["cuisine"], r["platform"], r["maps_url"]]
        for r in rows
    ]
    buf = build_excel(excel_rows, city)
    filename = f"choice_{city.lower().replace(' ', '_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    print("🚀 Open http://localhost:5000")
    app.run(debug=True, port=5000)
